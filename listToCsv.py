#! python3

import openpyxl
import csv
import sys

def main():
    rfile = 'テスト.xlsx' # ファイル名
    wb = openpyxl.load_workbook(rfile) # ファイルを開く
    sheetList = wb.get_sheet_names()
    sheet = wb.get_sheet_by_name(sheetList[0])
    
    wfile = 'shotaikyaku.csv'
    with open(wfile, 'w', encoding='utf-8') as fp:
        writer = csv.writer(fp)
        for r in range(2, sheet.max_row+1): # 一行目は飛ばす
            column = [str(sheet.cell(row=r,column=c).value or '') for c in range(1, sheet.max_column+1)]
            print(column)
            writer.writerow(column)

if __name__ == '__main__':
    main()
